import json, sqlite3
from dataclasses import dataclass
from pathlib import Path
from sqlalchemy import select
from langchain.agents import create_agent
from langchain.agents.structured_output import ToolStrategy
from langchain.agents.middleware import ModelCallLimitMiddleware, ToolCallLimitMiddleware, ModelRetryMiddleware, dynamic_prompt, ModelRequest
from langchain.tools import tool, ToolRuntime
from langchain_openai import ChatOpenAI
from langgraph.checkpoint.sqlite import SqliteSaver
from openpyxl import load_workbook
from .config import settings
from .database import SessionLocal
from .models import Document, DocumentChunk, Task, Artifact
from .schemas import TaskResult
from .security import safe_child
from .services import search_chunks, create_docx_artifact, create_xlsx_artifact

@dataclass
class AgentContext:
    workspace_id: str
    task_id: str

def _doc(db, context: AgentContext, document_id: str):
    d=db.get(Document,document_id)
    if not d or d.workspace_id!=context.workspace_id: raise ValueError("DOCUMENT_OUTSIDE_WORKSPACE")
    return d

@tool
def list_documents(runtime: ToolRuntime[AgentContext]) -> str:
    """列出当前工作区可用文档、格式、解析状态与元数据。"""
    with SessionLocal() as db:
        docs=db.scalars(select(Document).where(Document.workspace_id==runtime.context.workspace_id))
        return json.dumps([{"id":d.id,"name":d.original_name,"status":d.status,"media_type":d.media_type,"metadata":d.metadata_json} for d in docs],ensure_ascii=False)

@tool
def search_documents(query: str, limit: int, runtime: ToolRuntime[AgentContext]) -> str:
    """在当前工作区执行检索，返回最多 10 个带真实 chunk ID 与位置的候选。"""
    with SessionLocal() as db:
        hits=search_chunks(db,runtime.context.workspace_id,query,min(max(limit,1),10))
        return json.dumps([{"chunk_id":c.id,"document_id":d.id,"document_name":d.original_name,"location":c.location_json,"content":c.content[:1800]} for _,c,d in hits],ensure_ascii=False)

@tool
def read_document_section(chunk_id: str, runtime: ToolRuntime[AgentContext]) -> str:
    """按真实 chunk ID 读取当前工作区的一个文档片段。"""
    with SessionLocal() as db:
        c=db.get(DocumentChunk,chunk_id)
        if not c: raise ValueError("CHUNK_NOT_FOUND")
        d=_doc(db,runtime.context,c.document_id)
        return json.dumps({"chunk_id":c.id,"document_name":d.original_name,"location":c.location_json,"content":c.content},ensure_ascii=False)

@tool
def inspect_document_overview(document_id: str, runtime: ToolRuntime[AgentContext]) -> str:
    """读取一份文档的代表性片段，用于回答主题、内容概述和结构类问题。"""
    with SessionLocal() as db:
        d=_doc(db,runtime.context,document_id)
        chunks=list(db.scalars(select(DocumentChunk).where(DocumentChunk.document_id==d.id).order_by(DocumentChunk.chunk_index)))
        substantive=[c for c in chunks if len(c.content.strip())>=20]
        if not substantive:
            substantive=chunks
        sample_size=min(12,len(substantive))
        if sample_size<=1:
            sampled=substantive[:1]
        else:
            indexes=sorted({round(i*(len(substantive)-1)/(sample_size-1)) for i in range(sample_size)})
            sampled=[substantive[i] for i in indexes]
        return json.dumps({
            "document_id":d.id,
            "document_name":d.original_name,
            "chunk_count":len(chunks),
            "representative_sections":[{"chunk_id":c.id,"location":c.location_json,"content":c.content[:1200]} for c in sampled],
        },ensure_ascii=False)

@tool
def inspect_workbook(document_id: str, runtime: ToolRuntime[AgentContext]) -> str:
    """检查 XLSX 的工作表、维度、表头和公式数量，不计算公式。"""
    with SessionLocal() as db:
        d=_doc(db,runtime.context,document_id)
        if not d.original_name.lower().endswith('.xlsx'): raise ValueError("NOT_XLSX")
        path=safe_child(settings.app_data_dir,"workspaces",d.workspace_id,"uploads",d.stored_name)
        wb=load_workbook(path,read_only=True,data_only=False)
        result=[]
        for ws in wb.worksheets:
            header=[c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
            formulas=sum(1 for row in ws.iter_rows() for c in row if isinstance(c.value,str) and c.value.startswith('='))
            result.append({"sheet":ws.title,"rows":ws.max_row,"columns":ws.max_column,"header":header,"formulas":formulas})
        wb.close(); return json.dumps(result,ensure_ascii=False,default=str)

@tool
def read_worksheet_range(document_id: str, sheet: str, cell_range: str, runtime: ToolRuntime[AgentContext]) -> str:
    """读取 XLSX 的受限区域，最多 500 个单元格。"""
    with SessionLocal() as db:
        d=_doc(db,runtime.context,document_id); path=safe_child(settings.app_data_dir,"workspaces",d.workspace_id,"uploads",d.stored_name)
        wb=load_workbook(path,read_only=True,data_only=False)
        if sheet not in wb.sheetnames: raise ValueError("SHEET_NOT_FOUND")
        cells=wb[sheet][cell_range]
        if sum(len(row) for row in cells)>500: raise ValueError("RANGE_TOO_LARGE")
        values=[[c.value for c in row] for row in cells]; wb.close(); return json.dumps(values,ensure_ascii=False,default=str)

@tool
def analyze_table(document_id: str, sheet: str, operation: str, column: str, runtime: ToolRuntime[AgentContext]) -> str:
    """对白名单操作 count、sum、mean、min、max 执行基础表格统计。"""
    if operation not in {"count","sum","mean","min","max"}: raise ValueError("OPERATION_NOT_ALLOWED")
    with SessionLocal() as db:
        d=_doc(db,runtime.context,document_id); path=safe_child(settings.app_data_dir,"workspaces",d.workspace_id,"uploads",d.stored_name)
        wb=load_workbook(path,read_only=True,data_only=True); ws=wb[sheet]; headers=[c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
        if column not in headers: raise ValueError("COLUMN_NOT_FOUND")
        idx=headers.index(column); values=[r[idx].value for r in ws.iter_rows(min_row=2) if isinstance(r[idx].value,(int,float))]; wb.close()
        value=len(values) if operation=="count" else (sum(values) if operation=="sum" else sum(values)/len(values) if operation=="mean" and values else min(values) if operation=="min" and values else max(values) if operation=="max" and values else None)
        return json.dumps({"operation":operation,"column":column,"value":value,"count":len(values)},ensure_ascii=False)

@tool
def create_docx(title: str, sections: list[dict], runtime: ToolRuntime[AgentContext]) -> str:
    """从结构化章节创建并重新打开验证 DOCX；不会修改上传文件。"""
    with SessionLocal() as db:
        task=db.get(Task,runtime.context.task_id); a=create_docx_artifact(db,task,title,sections); return json.dumps({"artifact_id":a.id,"name":a.name})

@tool
def create_xlsx(sheets: list[dict], runtime: ToolRuntime[AgentContext]) -> str:
    """从结构化工作表定义创建并重新打开验证 XLSX；不会修改上传文件。"""
    with SessionLocal() as db:
        task=db.get(Task,runtime.context.task_id); a=create_xlsx_artifact(db,task,sheets); return json.dumps({"artifact_id":a.id,"name":a.name})

@tool
def validate_artifact(artifact_id: str, runtime: ToolRuntime[AgentContext]) -> str:
    """重新检查当前任务产物的路径、大小与可打开性。"""
    with SessionLocal() as db:
        a=db.get(Artifact,artifact_id)
        if not a or a.task_id!=runtime.context.task_id: raise ValueError("ARTIFACT_OUTSIDE_TASK")
        path=safe_child(settings.app_data_dir,*Path(a.path).parts)
        if a.type=='docx': from docx import Document as D; D(path)
        if a.type=='xlsx': load_workbook(path,read_only=True).close()
        return json.dumps({"valid":path.is_file() and path.stat().st_size>0,"size_bytes":path.stat().st_size})

@dynamic_prompt
def workspace_prompt(request: ModelRequest) -> str:
    return f"""你是 OfficeDocumentAgent。当前 workspace_id={request.runtime.context.workspace_id}，task_id={request.runtime.context.task_id}。
只能使用工具读取当前工作区 ready 文档；不得访问网络、系统命令、任意 SQL 或其他路径；不得修改上传文件。
关键事实必须引用工具真实返回的 chunk/位置；信息不足或矛盾必须写入 warnings；生成文件后必须调用 validate_artifact。
当用户询问整份文档讲了什么、主题、结构或概述时：先调用 list_documents，再对目标文档调用一次 inspect_document_overview；根据代表性片段直接作答，不要为了穷举全文而反复 search_documents/read_document_section。
最终输出 TaskResult，中文简洁，citation id 必须与检索结果一致。"""

def invoke_agent(task: Task) -> TaskResult:
    model=ChatOpenAI(model=settings.qwen_model,api_key=settings.dashscope_api_key,base_url=settings.dashscope_base_url,temperature=settings.qwen_temperature,timeout=settings.qwen_timeout_seconds,max_retries=0,extra_body={"enable_thinking":False})
    conn=sqlite3.connect(settings.app_data_dir/'checkpoints.db',check_same_thread=False)
    saver=SqliteSaver(conn)
    agent=create_agent(model=model,tools=[list_documents,search_documents,read_document_section,inspect_document_overview,inspect_workbook,read_worksheet_range,analyze_table,create_docx,create_xlsx,validate_artifact],middleware=[workspace_prompt,ModelCallLimitMiddleware(run_limit=20,exit_behavior='error'),ToolCallLimitMiddleware(run_limit=40,exit_behavior='error'),ModelRetryMiddleware(max_retries=settings.qwen_max_retries,backoff_factor=2.0,on_failure='error')],response_format=ToolStrategy(TaskResult,handle_errors=lambda e:f"结构不合法，请修复；最多两次：{type(e).__name__}"),context_schema=AgentContext,checkpointer=saver)
    try:
        state=agent.invoke({"messages":[{"role":"user","content":task.user_message}]},context=AgentContext(task.workspace_id,task.id),config={"configurable":{"thread_id":task.thread_id}})
        structured=state.get("structured_response")
        if structured is None:
            messages=state.get("messages",[])
            last_content=str(getattr(messages[-1],"content",""))[:500] if messages else ""
            raise RuntimeError(f"MODEL_STRUCTURED_OUTPUT_MISSING: {last_content or '模型未返回结构化结果'}")
        return structured
    finally: conn.close()
